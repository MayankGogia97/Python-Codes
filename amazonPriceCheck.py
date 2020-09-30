import os
import pyperclip
import openpyxl
import sys
from bs4 import BeautifulSoup
import requests
import datetime as dt

headers = {'user-agent': 'Chrome'}          # User agents so that we can web scrape without amazon tagging as bot
center = openpyxl.styles.Alignment(horizontal='center', vertical='center')      # center alingment attribute

# Function to update the prices


def updatePrices():

    stamp = dt.date.today().isoformat()         # The column heddings to track progress

    if ws.max_column == 12:                     # Trimming prices older than 10 iterations
        print('Trimming old price columns...')
        moveString = 'D3:L' + str(ws.max_row)
        ws.move_range(moveString, cols=-1)
        ws.unmerge_cells('C1:L1')
        ws.merge_cells('C1:K1')
        ws['C1'].alignment = center
        ws.delete_cols(12)

    if ws['C3'].value is None and ws.max_column == 3:                   # If no prices are there
        colNum = 3
    else:
        colNum = ws.max_column + 1

    ws.cell(row=2, column=colNum).value = stamp
    ws.cell(row=2, column=colNum).alignment = center
    ws.column_dimensions[openpyxl.utils.get_column_letter(colNum)].width = 11
    for rowNum in range(3, ws.max_row + 1):

        link = ws.cell(row=rowNum, column=2).value[12:-7]
        res = requests.get(link, headers=headers)
        res.raise_for_status()
        soup = BeautifulSoup(res.text, 'lxml')

        try:
            priceText = soup.find('span', {'id': 'priceblock_dealprice'}).getText()

        except:  # noqa: E722
            priceText = soup.find('span', {'id': 'priceblock_ourprice'}).getText()

        try:
            price = int(priceText[2:-3].replace(',', ''))

        except:  # noqa: E722
            price = int(priceText[2:-3])

        ws.cell(row=rowNum, column=colNum).value = price
        ws.cell(row=rowNum, column=colNum).alignment = center
    if not ws.max_column == 3:
        if not ws.max_column == 4:
            unmergeString = 'C1:' + openpyxl.utils.get_column_letter(ws.max_column-1) + '1'
            ws.unmerge_cells(unmergeString)

        mergeString = 'C1:' + openpyxl.utils.get_column_letter(ws.max_column) + '1'
        ws.merge_cells(mergeString)
        ws['C1'].alignment = center

    print('Prices updated...')


def updateWidth():                  # Update the width of the name column along the entries

    rowWidths = []
    for rowNum in range(3, ws.max_row + 1):
        rowWidths.append(len(ws.cell(row=rowNum, column=1).value))
    maxWidth = max(rowWidths)
    if maxWidth > 65:
        maxWidth = 65
    ws.column_dimensions['A'].width = maxWidth
    print('Width of column adjusted...')


def reset():                    # Reset the Excel File

    if ws.max_column > 3:
        unmergeString = 'C1:' + openpyxl.utils.get_column_letter(ws.max_column) + '1'
        ws.unmerge_cells(unmergeString)
        ws['C1'].alignment = center

    for colNum in range(3, ws.max_column+1):

        ws.cell(row=2, column=colNum).value = None

    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row)

    ws.column_dimensions['A'].width = ws.column_dimensions['B'].width = ws.column_dimensions['C'].width = 12
    print('The Excel File has been resetted')


def addLink(link):                      # Add new links to the file

    res = requests.get(link, headers=headers)
    soup = BeautifulSoup(res.text, 'lxml')
    rowIndex = ws.max_row + 1
    name = soup.find('span', {'id': 'productTitle'}).getText().strip()
    ws.cell(row=rowIndex, column=1).value = name
    ws.cell(row=rowIndex, column=2).value = '=HYPERLINK("' + link + '","...")'
    ws.cell(row=rowIndex, column=2).style = 'Hyperlink'
    ws.cell(row=rowIndex, column=2).alignment = center
    print('Item added successfully...')
    updatePrices()                      # Call to update all the prices
    updateWidth()                       # Call to update the name column's width


def delItem(name):                      # Delete the item whose name is passed

    for rowNum in range(3, ws.max_row+1):

        if ws.cell(row=rowNum, column=1).value == name:

            break

    else:

        print('No item has that name.\nExiting without deleting anything...')
        sys.exit()

    rowLimit = ws.max_row
    if rowLimit > 3:
        if rowNum < ws.max_row:
            moveString = 'A' + str(rowNum + 1) + ':' + openpyxl.utils.get_column_letter(ws.max_column) + str(rowLimit)
            ws.move_range(moveString, rows=-1)
    ws.delete_rows(rowLimit)
    print('Item deleted...')

    if ws['A3'].value is None:
        print('Since no items were left, the excel file has to be reset...')
        reset()


os.chdir(r'G:\Codes\Python')
wb = openpyxl.load_workbook('amazonPriceHistory.xlsx')
wb.save('amazonPriceHistoryBackup.xlsx')
wb = openpyxl.load_workbook('amazonPriceHistory.xlsx')
ws = wb.active

if len(sys.argv) > 1:

    if sys.argv[1] == 'add':

        searchLink = pyperclip.paste()
        addLink(searchLink)

    elif sys.argv[1] == 'del':
        itemName = pyperclip.paste()
        delItem(itemName)

    elif sys.argv[1] == 'reset':
        reset()

else:

    if ws.max_row == 2:
        print('There are currently no items in the list to update.\nExiting...')
        sys.exit()

    updatePrices()

ws.freeze_panes = 'A3'
wb.save('amazonPriceHistory.xlsx')
print('Done\nDo you want to open the file(Y/y)?\n')
choice = input()
if choice in ['y', 'Y']:
    os.startfile('amazonPriceHistory.xlsx')
